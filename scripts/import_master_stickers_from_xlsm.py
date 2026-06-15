"""
import_master_stickers_from_xlsm.py

Genera data/stickers_master.csv a partir de la grilla del album en
../../docs/StickerAlbumWC2026.xlsm (hoja "Stickers").

Fase 0.5: cambio de criterio explicito - todavia NO se importan nombres de
jugadores reales (hay varias fuentes posibles y no se quiere depender de
ninguna hasta confirmar la oficial). `player_name_or_detail` se genera de
forma generica a partir del codigo:

    MEX1  -> "Mexico 1"
    ARG10 -> "Argentina 10"
    FWC1  -> "Especial FIFA 1"
    00    -> "Logo Oficial Mundial 2026"

Estructura de la hoja "Stickers":
    - Fila 7: "Specials" | "Fifa" | "00" | FW1 | FW2 | ... | FW19
    - Fila 8: encabezado de columnas (numeros 1..20)
    - Filas 9..56: 12 grupos (A-L) x 4 paises, cada fila = un pais con sus
      20 codigos de figuritas (col C..V), columna A = "Group X" solo en la
      primera fila de cada grupo.

Reglas aplicadas (ver IMPORT_MASTER_STICKERS_REPORT.md para el detalle):
    - code/country_code/number se derivan de los codigos de cada fila;
      country_code = prefijo dominante de la fila (corrige typos como
      "Che16" -> "CZE16").
    - Especiales FIFA (prefijo FWC): country_code="FWC",
      country_name="FIFA World Cup 2026", type="special". `group` se deja
      vacio para que /api/reports/album los siga agrupando en `special`
      (mismo comportamiento que el dataset de prueba anterior).
    - Celdas que no son codigos de figurita ni "00" (ej. encabezados) se
      ignoran y quedan registradas en el reporte.
    - El codigo "00" es el logo/portada/sticker inicial del album: se
      importa como type="logo", country_code="FWC", number=0,
      player_name_or_detail="Logo Oficial Mundial 2026".
    - Codigos duplicados se ignoran (se conserva la primera aparicion).

Requiere: openpyxl (pip install openpyxl).

Uso:
    python scripts/import_master_stickers_from_xlsm.py
"""

import csv
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent  # stickercontrol-2026/
REPO_ROOT = PROJECT_ROOT.parent  # ALBUM PANINI/

XLSM_PATH = REPO_ROOT / "docs" / "StickerAlbumWC2026.xlsm"
OUTPUT_CSV = PROJECT_ROOT / "data" / "stickers_master.csv"
REPORT_PATH = PROJECT_ROOT / "IMPORT_MASTER_STICKERS_REPORT.md"

SHEET_NAME = "Stickers"
FIRST_DATA_ROW = 7
FIRST_DATA_COL = 3  # columna C
LAST_DATA_COL = 22  # columna V

CODE_RE = re.compile(r"^([A-Za-z]{2,4})(\d{1,2})$")
GROUP_RE = re.compile(r"^Group\s+([A-L])$", re.IGNORECASE)

# Codigo especial "00": logo/portada/sticker inicial del album. No matchea
# CODE_RE (no tiene prefijo de letras), por eso se trata aparte.
LOGO_CODE = "00"
LOGO_ROW = {
    "code": LOGO_CODE,
    "group": "",
    "country_code": "FWC",
    "country_name": "FIFA World Cup 2026",
    "number": 0,
    "player_name_or_detail": "Logo Oficial Mundial 2026",
    "type": "logo",
}

CSV_COLUMNS = [
    "code",
    "group",
    "country_code",
    "country_name",
    "number",
    "player_name_or_detail",
    "type",
]


def extract_rows(ws):
    rows = []
    seen_codes: set[str] = set()
    duplicates: list[tuple[int, int, str]] = []
    ignored_cells: list[tuple[int, int, str]] = []
    corrections: list[tuple[int, int, str, str, str]] = []
    special_codes: list[tuple[str, str]] = []
    current_group: str | None = None

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        col_a = ws.cell(row=r, column=1).value
        col_b = ws.cell(row=r, column=2).value

        if col_b is None:
            continue
        col_b = str(col_b).strip()
        if col_b == "" or col_b.lower() == "country":
            continue  # fila vacia o de encabezado

        is_special = col_b.lower() == "fifa"

        if isinstance(col_a, str):
            m = GROUP_RE.match(col_a.strip())
            if m:
                current_group = m.group(1).upper()

        if is_special:
            group = ""
            country_name = "FIFA World Cup 2026"
        else:
            group = current_group or ""
            country_name = col_b

        # Codigos de figurita validos en esta fila (col C..V)
        cells: list[tuple[int, str, str, int]] = []
        for c in range(FIRST_DATA_COL, LAST_DATA_COL + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            raw = str(val).strip()
            if raw == "":
                continue

            if raw == LOGO_CODE:
                if LOGO_CODE in seen_codes:
                    duplicates.append((r, c, LOGO_CODE))
                    continue
                seen_codes.add(LOGO_CODE)
                rows.append(dict(LOGO_ROW))
                special_codes.append((LOGO_CODE, LOGO_ROW["player_name_or_detail"]))
                continue

            m = CODE_RE.match(raw)
            if not m:
                ignored_cells.append((r, c, raw))
                continue
            prefix = m.group(1).upper()
            number = int(m.group(2))
            cells.append((c, raw, prefix, number))

        if not cells:
            continue

        if is_special:
            dominant = "FWC"
            country_code = "FWC"
        else:
            dominant = Counter(prefix for _, _, prefix, _ in cells).most_common(1)[0][0]
            country_code = dominant

        for c, raw_code, prefix, number in cells:
            code = f"{dominant}{number}"
            if code != raw_code:
                corrections.append((r, c, raw_code, code, country_name))

            if code in seen_codes:
                duplicates.append((r, c, code))
                continue
            seen_codes.add(code)

            if is_special:
                player = f"Especial FIFA {number}"
                sticker_type = "special"
            else:
                player = f"{country_name} {number}"
                sticker_type = "standard"

            rows.append(
                {
                    "code": code,
                    "group": group,
                    "country_code": country_code,
                    "country_name": country_name,
                    "number": number,
                    "player_name_or_detail": player,
                    "type": sticker_type,
                }
            )

    return rows, duplicates, ignored_cells, corrections, special_codes


def write_csv(rows: list[dict]) -> None:
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_report(rows, duplicates, ignored_cells, corrections, special_codes) -> None:
    by_group: Counter[str] = Counter()
    by_country: Counter[str] = Counter()
    for row in rows:
        by_group[row["group"] or "(especiales)"] += 1
        by_country[row["country_code"]] += 1

    lines = []
    lines.append("# Reporte de importacion - master_stickers")
    lines.append("")
    lines.append(f"- Archivo origen: `docs/StickerAlbumWC2026.xlsm`")
    lines.append(f"- Hoja usada: `{SHEET_NAME}`")
    lines.append(f"- CSV generado: `data/stickers_master.csv`")
    lines.append(f"- Codigos detectados (total): **{len(rows)}**")
    lines.append("")

    lines.append("## Cantidad por grupo")
    lines.append("")
    lines.append("| Grupo | Cantidad |")
    lines.append("|---|---|")
    for group in sorted(by_group, key=lambda g: (g == "(especiales)", g)):
        lines.append(f"| {group} | {by_group[group]} |")
    lines.append("")

    lines.append("## Cantidad por pais (country_code)")
    lines.append("")
    lines.append("| country_code | Cantidad |")
    lines.append("|---|---|")
    for country_code in sorted(by_country):
        lines.append(f"| {country_code} | {by_country[country_code]} |")
    lines.append("")

    lines.append("## Codigos duplicados ignorados")
    lines.append("")
    if duplicates:
        lines.append("| Fila | Columna | Codigo |")
        lines.append("|---|---|---|")
        for r, c, code in duplicates:
            lines.append(f"| {r} | {c} | {code} |")
    else:
        lines.append("Ninguno.")
    lines.append("")

    lines.append("## Celdas ignoradas (no son codigos de figurita)")
    lines.append("")
    if ignored_cells:
        lines.append("| Fila | Columna | Valor |")
        lines.append("|---|---|---|")
        for r, c, raw in ignored_cells:
            lines.append(f"| {r} | {c} | `{raw}` |")
    else:
        lines.append("Ninguna.")
    lines.append("")

    lines.append("## Filas dudosas / corregidas")
    lines.append("")
    if corrections:
        lines.append("| Fila | Columna | Valor original | Codigo corregido | Pais |")
        lines.append("|---|---|---|---|---|")
        for r, c, raw_code, code, country_name in corrections:
            lines.append(f"| {r} | {c} | `{raw_code}` | `{code}` | {country_name} |")
        lines.append("")
        lines.append(
            "El codigo se corrige al prefijo dominante de la fila (el codigo "
            "de pais que comparten el resto de las celdas), para que el pais "
            "quede con sus 20 figuritas sin huecos."
        )
    else:
        lines.append("Ninguna.")
    lines.append("")

    lines.append("## Codigos especiales importados")
    lines.append("")
    if special_codes:
        for code, player in special_codes:
            lines.append(f"- `{code}` -> {player}")
    else:
        lines.append("Ninguno.")
    lines.append("")

    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    if not XLSM_PATH.exists():
        print(f"ERROR: no existe {XLSM_PATH}")
        return 1

    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: la hoja '{SHEET_NAME}' no existe. Hojas disponibles: {wb.sheetnames}")
        return 1

    ws = wb[SHEET_NAME]
    rows, duplicates, ignored_cells, corrections, special_codes = extract_rows(ws)

    if not rows:
        print("ERROR: no se detecto ningun codigo de figurita.")
        return 1

    write_csv(rows)
    write_report(rows, duplicates, ignored_cells, corrections, special_codes)

    print(f"OK: {len(rows)} figuritas escritas en {OUTPUT_CSV}")
    print(f"Duplicados ignorados: {len(duplicates)}")
    print(f"Celdas ignoradas: {len(ignored_cells)}")
    print(f"Filas corregidas: {len(corrections)}")
    print(f"Codigos especiales importados: {len(special_codes)}")
    print(f"Reporte: {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
